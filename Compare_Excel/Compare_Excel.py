#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import tkinter as tk
import numpy as np
import openpyxl
import time
from tkinter import *
from tkinter import filedialog
from openpyxl import Workbook 
from openpyxl.styles import PatternFill 


def change_cell_color(file_path, row_array, column_array):
    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Create a PatternFill object with the desired color
    fill = PatternFill(start_color="00FFFF00", fill_type="solid")

    # Loop through the specified rows and columns and apply the color to each cell
    for item in zip(row_array,column_array):
        cell = sheet.cell(row=item[0]+2, column=item[1]+1)
        cell.fill = fill
   # Save the changes to the Excel file
    workbook.save(file_path)
    
def compare_excel_files():
    # Ask the user to select the two Excel files
    file1_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    file2_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    destination = 'C:\\Users\\Administrator\\Desktop\\export_dataframe.xlsx'
    
    # Check if the user selected two valid files
    if not file1_path or not file2_path:
        return

    # Load the Excel files into dataframes
    df1 = pd.read_excel(file1_path)
    df2 = pd.read_excel(file2_path)

    # Find the differences between the two dataframes
    comparison_values = df1.values == df2.values
    rows,cols=np.where(comparison_values==False)
    
    
    for item in zip(rows,cols):
        df1.iloc[item[0], item[1]] = '{} --> {}'.format(df1.iloc[item[0], item[1]],df2.iloc[item[0], item[1]])
  
 
    df1.to_excel(destination, index=False)
    time.sleep(5)
    change_cell_color(destination, rows, cols)
    
    Output.insert(END, "Succees!")


# Create a Tkinter application window
root = Tk()
root.title("Excel File Comparator")

# Add a button to trigger the comparison process
compare_button = tk.Button(root, text="Compare Excel Files", command=compare_excel_files)
compare_button.pack(pady=10)

Output = Text(root, height = 5,
              width = 25,
              bg = "light cyan")
Output.pack()

# Run the Tkinter main loop
root.mainloop()

